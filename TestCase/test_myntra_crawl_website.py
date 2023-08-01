import os
import openpyxl
from utils import get_hash_key, get_products, get_total_pages

# Check if the data file already exists?
file_name = './data.xlsx'
hash = {}
if os.path.isfile(file_name):
    print(f"{file_name} exists! Reading items from {file_name}!")
    # If the file already exists, then load all the items into the hash.
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        id, name, *_ = row
        hash_key = get_hash_key(id, name)
        if hash_key not in hash:
            hash[hash_key] = row
        else:
            print(f"{file_name} is corrupt! Duplicate entries found, please delete the file and re-run the script.")
            print(f'Duplicate item: {row}')
            exit()
else:
    print(f"{file_name} does not exist, creating a new file!")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    field_names = ["productId", "productName", "price", "rating", "ratingCount", "searchImage", "landingPageUrl"]
    sheet.append(field_names)

entries_processed = 0  # Counter to keep track of processed entries
total_pages = get_total_pages()
for i in range(1, total_pages + 1):
    if entries_processed >= 50:
        break  # Stop processing further when 50 entries are reached
    products = get_products(i)
    for product in products:
        entries_processed += 1
        productId = product['productId']
        landingPageUrl = product['landingPageUrl']
        productName = product['productName']
        price = product['price']
        rating = product['rating']
        searchImage = product['searchImage']
        ratingCount = product['ratingCount']
        hash_key = get_hash_key(productId, productName)
        if hash_key not in hash:
            print(f"Unique product found! {productName}.")
            row_data = [productId, productName, price, rating, ratingCount, searchImage, landingPageUrl]
            sheet.append(row_data)
            hash[hash_key] = row_data

workbook.save(file_name)
workbook.close()
